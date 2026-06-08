#!/usr/bin/env python3
"""
Converte o PCASP Estendido do TCE-PR (XLSX oficial do Tribunal de Contas do
Estado do Paraná) para o CSV de importação do Gênesis (formato consumido por
src/services/importador-plano-contas.ts).

Origem: "PLANO DE CONTAS PARANÁ_2026 - Estendido" — aba PCASP-TCE-PR-2026.
Diferente do XLSX da STN (ver xlsx_para_csv_pcasp.py): aqui o código tem 12
segmentos e a árvore chega ao nível 9; o flag de movimento vem de tpEscrituracao.

Layout da aba (cabeçalho na linha 5, dados a partir da linha 6):
  - Col N  "Conta"                 → codigo          (ex.: 1.1.2.1.1.70.01.01.00.00.00.00)
  - Col P  "dsConta"               → descricao
  - Col R  "tpEscrituracao"        → admiteMovimento (== 'S')
  - Col Q  "tpNaturezaSaldo"       → naturezaSaldo       (D/C/X bruto)
  - Col S  "tpNaturezaInformacao"  → naturezaInformacao  (P/O/C bruto)
  - Col T  "tpSuperavitFinanceiro" → superavitFinanceiro (F/P/X/O bruto)
  - Col V  "Função"                → funcao              (texto; o mapeamento dos
    códigos brutos para os enums é feito no importador TS)
  - codigoPai derivado zerando o último segmento não-zero do código
"""

from __future__ import annotations
import csv
import sys
from pathlib import Path
from openpyxl import load_workbook

SHEET = "PCASP-TCE-PR-2026"
COL_CONTA = 13          # N
COL_DSCONTA = 15        # P
COL_ESCRITURACAO = 17   # R
COL_NAT_SALDO = 16      # Q  tpNaturezaSaldo
COL_NAT_INFO = 18       # S  tpNaturezaInformacao
COL_SUPERAVIT = 19      # T  tpSuperavitFinanceiro
COL_FUNCAO = 21         # V  Função
HEADER_ROW = 5          # dados começam na linha seguinte


def parent_code(codigo: str) -> str:
    """1.1.2.1.1.70.01.01... → zera o último segmento não-zero (preserva a largura)."""
    segs = codigo.split(".")
    for i in range(len(segs) - 1, -1, -1):
        if segs[i].strip("0") != "":
            if i == 0:
                return ""
            zero = "0" * len(segs[i])
            return ".".join(segs[:i] + [zero] + segs[i + 1 :])
    return ""


def converter(xlsx: Path, csv_saida: Path) -> None:
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[SHEET]

    # codigo → (descricao, admiteMovimento, natSaldo, natInfo, superavit, funcao)
    contas: dict[str, tuple[str, bool, str, str, str, str]] = {}

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < HEADER_ROW:  # pula título/cabeçalho (linhas 1..5, 0-based < 5)
            continue
        conta = row[COL_CONTA]
        if conta is None or not str(conta).strip():
            continue
        codigo = str(conta).strip()
        # dsConta pode ter \n embutido; o parser do importador faz split por
        # linha sem respeitar quoting, então colapsamos espaços/quebras.
        titulo = " ".join(str(row[COL_DSCONTA] or "").split()).strip()
        if not titulo:
            print(f"AVISO: conta {codigo} sem descrição — pulada.", file=sys.stderr)
            continue
        admite = str(row[COL_ESCRITURACAO] or "").strip().upper() == "S"
        nat_saldo = str(row[COL_NAT_SALDO] or "").strip().upper()
        nat_info = str(row[COL_NAT_INFO] or "").strip().upper()
        superavit = str(row[COL_SUPERAVIT] or "").strip().upper()
        # Função: texto longo, colapsa quebras/espaços (mesma razão do título).
        funcao = " ".join(str(row[COL_FUNCAO] or "").split()).strip()
        contas[codigo] = (titulo, admite, nat_saldo, nat_info, superavit, funcao)

    # Validação: todo pai referenciado deve existir no próprio arquivo.
    faltando = [(c, parent_code(c)) for c in contas if parent_code(c) and parent_code(c) not in contas]
    if faltando:
        print(f"ERRO: {len(faltando)} contas com pai ausente:", file=sys.stderr)
        for c, p in faltando[:10]:
            print(f"  {c} → {p}", file=sys.stderr)
        sys.exit(1)

    with csv_saida.open("w", newline="", encoding="utf-8") as fp:
        w = csv.writer(fp)
        w.writerow([
            "codigo", "descricao", "codigoPai", "admiteMovimento",
            "naturezaSaldo", "naturezaInformacao", "superavitFinanceiro", "funcao",
        ])
        for codigo in sorted(contas, key=lambda c: [int(s) for s in c.split(".")]):
            descricao, admite, nat_saldo, nat_info, superavit, funcao = contas[codigo]
            w.writerow([
                codigo, descricao, parent_code(codigo), "true" if admite else "false",
                nat_saldo, nat_info, superavit, funcao,
            ])

    print(f"OK: {len(contas)} contas em {csv_saida}")


if __name__ == "__main__":
    entrada = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(
        "data/PLANO DE CONTAS PARANÁ_2026 - Estendido - Versao 1.0d - Publicada em 24-03-2026.xlsx"
    )
    saida = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("data/pcasp_estendido_2026.csv")
    converter(entrada, saida)
