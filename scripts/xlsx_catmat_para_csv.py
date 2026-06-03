#!/usr/bin/env python3
"""
Converte o Catálogo de Materiais (CATMAT) do XLSX oficial para o CSV de
importação do Gênesis (consumido por scripts/importar_catmat_2026.ts).

Fonte: "Novo Catálogo de Materiais ..." — aba "Materiais".
  - Col G "Código do Item"     → codigo
  - Col H "Descrição do Item"  → descricao  (whitespace colapsado)

O tipo (MATERIAL) e a unidade de medida ('UN') são constantes aplicadas na
importação — a planilha do CATMAT não traz coluna de unidade (decisão do Marco).

Uso: xlsx_catmat_para_csv.py [<xlsx>] [<saida.csv>]
"""

from __future__ import annotations
import csv
import sys
from pathlib import Path
from openpyxl import load_workbook

SHEET = "Materiais"
COL_CODIGO = 6       # G
COL_DESCRICAO = 7    # H
HEADER_ROWS = 1      # dados a partir do índice 0-based >= 1 (linha 2)


def normaliza_codigo(v) -> str:
    """O código CATMAT é inteiro, mas o openpyxl lê algumas células como float
    (ex.: 431794.0). str() direto geraria "431794.0" e duplicaria o item.
    Coage int/float para a forma inteira canônica; texto fica como está."""
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else str(v)
    return str(v).strip()


def converter(xlsx: Path, saida: Path) -> None:
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[SHEET]

    itens: dict[str, str] = {}  # codigo → descricao (primeiro vence; CATMAT é único por item)
    dups = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < HEADER_ROWS:
            continue
        cod = row[COL_CODIGO] if len(row) > COL_CODIGO else None
        if cod is None or not str(cod).strip():
            continue
        codigo = normaliza_codigo(cod)
        descricao = " ".join(str(row[COL_DESCRICAO] or "").split()).strip()
        if not descricao:
            print(f"AVISO: item {codigo} sem descrição — pulado.", file=sys.stderr)
            continue
        if codigo in itens:
            dups += 1
            continue
        itens[codigo] = descricao

    if dups:
        print(f"AVISO: {dups} código(s) duplicado(s) no arquivo — mantida a 1ª ocorrência.", file=sys.stderr)

    with saida.open("w", newline="", encoding="utf-8") as fp:
        w = csv.writer(fp)
        w.writerow(["codigo", "descricao"])
        for codigo in sorted(itens):
            w.writerow([codigo, itens[codigo]])

    print(f"OK: {len(itens)} itens em {saida}")


if __name__ == "__main__":
    entrada = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("data/Novo Catálogo de Materiais 26-05-26.xlsx")
    saida = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("data/catmat_2026.csv")
    converter(entrada, saida)
