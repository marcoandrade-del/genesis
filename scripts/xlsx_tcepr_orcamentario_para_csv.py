#!/usr/bin/env python3
"""
Converte os planos ORÇAMENTÁRIOS do TCE-PR (Receita / Despesa) para o CSV de
importação do Gênesis (formato consumido por importador-plano-receita.ts /
importador-plano-despesa.ts — mesmas 4 colunas do contábil).

Diferente do contábil (ver xlsx_tcepr_para_csv.py), nestas planilhas o flag de
movimento vem da coluna "Nível (S/A)" / idTipoNivelConta: A = Analítica (folha,
admite movimento), S = Sintética (agregadora).

Parametrizado por planilha/colunas (0-based), pois receita e despesa diferem:
  RECEITA  aba "Plano Receita 2026":      Código=N(13)  dsDesdobramento=P(15)  idTipoNivelConta=R(17)
  DESPESA  aba "PC Desp-2026 1.0e-3902":  CÓDIGO=H(7)   TÍTULO=J(9)            Nível(S/A)=K(10)

Cabeçalho na linha 7 (1-based); dados a partir da linha 8.

Uso: xlsx_tcepr_orcamentario_para_csv.py <xlsx> <aba> <col_codigo> <col_desc> <col_sa> <saida.csv>
"""

from __future__ import annotations
import csv
import sys
from pathlib import Path
from openpyxl import load_workbook

HEADER_ROWS = 7  # dados começam no índice 0-based >= 7 (linha 8)


def parent_code(codigo: str) -> str:
    """Zera o último segmento não-zero, preservando a largura de cada segmento."""
    segs = codigo.split(".")
    for i in range(len(segs) - 1, -1, -1):
        if segs[i].strip("0") != "":
            if i == 0:
                return ""
            return ".".join(segs[:i] + ["0" * len(segs[i])] + segs[i + 1 :])
    return ""


def converter(xlsx: Path, sheet: str, c_cod: int, c_desc: int, c_sa: int, saida: Path) -> None:
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[sheet]

    contas: dict[str, tuple[str, bool]] = {}  # codigo → (descricao, admiteMovimento)
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < HEADER_ROWS:
            continue
        cod = row[c_cod]
        if cod is None or not str(cod).strip():
            continue
        codigo = str(cod).strip()
        titulo = " ".join(str(row[c_desc] or "").split()).strip()
        if not titulo:
            print(f"AVISO: conta {codigo} sem descrição — pulada.", file=sys.stderr)
            continue
        admite = str(row[c_sa] or "").strip().upper() == "A"
        contas[codigo] = (titulo, admite)

    faltando = [(c, parent_code(c)) for c in contas if parent_code(c) and parent_code(c) not in contas]
    if faltando:
        print(f"ERRO: {len(faltando)} contas com pai ausente:", file=sys.stderr)
        for c, p in faltando[:10]:
            print(f"  {c} → {p}", file=sys.stderr)
        sys.exit(1)

    with saida.open("w", newline="", encoding="utf-8") as fp:
        w = csv.writer(fp)
        w.writerow(["codigo", "descricao", "codigoPai", "admiteMovimento"])
        for codigo in sorted(contas, key=lambda c: [int(s) for s in c.split(".")]):
            descricao, admite = contas[codigo]
            w.writerow([codigo, descricao, parent_code(codigo), "true" if admite else "false"])

    print(f"OK: {len(contas)} contas em {saida}")


if __name__ == "__main__":
    if len(sys.argv) != 7:
        print(__doc__, file=sys.stderr)
        sys.exit(2)
    converter(Path(sys.argv[1]), sys.argv[2], int(sys.argv[3]), int(sys.argv[4]), int(sys.argv[5]), Path(sys.argv[6]))
