#!/usr/bin/env python3
"""
Converte o PCASP Estendido oficial (XLSX da STN) para o CSV de importação
do Gênesis (formato consumido por src/services/importador-plano-contas.ts).

Origem: https://thot-arquivos.tesouro.gov.br/publicacao/47458
        (PCASP Estendido 2024, página Tesouro Transparente)

Mapeamento:
  - Col H "Conta"               → codigo                (ex.: 1.1.1.1.1.01.00)
  - Col I "Título"              → descricao             (capitalização do XLSX)
  - codigoPai derivado zerando o último segmento não-zero do código
  - Col N "Nível de detalhamento" == "Último" → admiteMovimento=true
  - Filtra Status == "Ativa"
"""

from __future__ import annotations
import csv
import sys
from pathlib import Path
from openpyxl import load_workbook


def parent_code(codigo: str) -> str:
    """1.1.1.1.1.01.01 → 1.1.1.1.1.01.00 → 1.1.1.1.1.00.00 → ... → ''."""
    segs = codigo.split(".")
    for i in range(len(segs) - 1, -1, -1):
        if segs[i] not in ("0", "00"):
            # Zera esse segmento e tudo à direita; corta zeros à direita não.
            largura = len(segs[i])
            zero = "0" * largura
            novos = segs[:i] + [zero] + segs[i + 1 :]
            # Se zerou o primeiro segmento, é raiz.
            if i == 0:
                return ""
            return ".".join(novos)
    return ""


def converter(xlsx: Path, csv_saida: Path) -> None:
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["PCASP 2024"]

    contas: dict[str, tuple[str, bool]] = {}  # codigo → (descricao, admiteMovimento)

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not row or not row[7]:
            continue
        codigo = str(row[7]).strip()
        # Algumas descrições têm \n embutido; o parser do importador faz
        # split(/\r?\n/) sem respeitar quoting, então colapsamos pra espaço.
        titulo = " ".join((row[8] or "").split()).strip()
        status = (row[12] or "").strip()
        nivel_det = (row[13] or "").strip()
        if status != "Ativa":
            continue
        if not titulo:
            continue
        admite = nivel_det == "Último"
        contas[codigo] = (titulo, admite)

    # Validação: todo pai referenciado deve existir.
    faltando: list[tuple[str, str]] = []
    for codigo in contas:
        pai = parent_code(codigo)
        if pai and pai not in contas:
            faltando.append((codigo, pai))

    if faltando:
        print(f"AVISO: {len(faltando)} contas com pai ausente:", file=sys.stderr)
        for c, p in faltando[:10]:
            print(f"  {c} → {p}", file=sys.stderr)
        if len(faltando) > 10:
            print(f"  ... e mais {len(faltando) - 10}", file=sys.stderr)
        sys.exit(1)

    # Escreve em ordem de código.
    with csv_saida.open("w", newline="", encoding="utf-8") as fp:
        w = csv.writer(fp)
        w.writerow(["codigo", "descricao", "codigoPai", "admiteMovimento"])
        for codigo in sorted(contas, key=lambda c: [int(s) for s in c.split(".")]):
            descricao, admite = contas[codigo]
            w.writerow([codigo, descricao, parent_code(codigo), "true" if admite else "false"])

    print(f"OK: {len(contas)} contas em {csv_saida}")


if __name__ == "__main__":
    # Caminhos via argumento; default mantém a fonte padrão em data/.
    # Saída default = mesmo nome do XLSX com extensão .csv.
    entrada = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("data/pcasp_estendido_2024.xlsx")
    saida = Path(sys.argv[2]) if len(sys.argv) > 2 else entrada.with_suffix(".csv")
    converter(entrada, saida)
